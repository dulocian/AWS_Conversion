import os
import csv
import string
import argparse
from openpyxl import load_workbook

def load_excel(file_path):
    """
    Load an Excel file and return the data and headers.

    :param file_path: Path to the Excel file.
    :return: Tuple of data (list of rows) and headers (list of column names).
    """
    headers = []
    data = []

    # Load the workbook and select the active worksheet
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=1, max_col=10, values_only=False):
        row_data = [cell.value for cell in row]  # List comprehension for row data
        if not headers:  # Separate headers from the rows data
            headers = row_data
        else:
            data.append(row_data)

    wb.close()  # Close the workbook when done
    return data, headers  # Return data excluding the headers

def flatten_data(lines, headers):
    """
    Flatten the data from the Excel sheet and track highest counts per alphabetical character.

    :param lines: Rows of data from the Excel sheet.
    :param headers: Column headers.
    :return: Tuple of flattened data and highest counts.
    """
    # Alphabet list to track cnt_alpha (A, B, C, etc.)
    alphabet = list(string.ascii_uppercase)

    # Initialize counters and data structures
    cnt_alpha = 0  # Tracks "lemma" count (A, B, C, ...)
    cnt_num = 0  # Tracks row count within each lemma (1, 2, 3, ...)
    cur_id = None  # To store the current entry ID

    flat_data = []  # Will hold the final flattened output
    entry = {}  # Stores temporary entry data

    highest = {}  # Track the highest count per alphabetical char

    for row in lines:
        # A new entry (ID) resets all counters and saves the previous entry
        if row[0] is not None:
            if cur_id is not None:
                entry["ID"] = cur_id
                flat_data.append(entry)

            # Reset for the new entry
            cnt_alpha = 0
            cnt_num = 0
            cur_id = row[0]  # Update the current ID
            entry = {}  # Reset the entry for the new ID

        # Each new lemma (first column after ID) increments the alphabetical counter
        if row[1] is not None:
            cnt_alpha += 1
            cnt_num = 0  # Reset row counter for new lemma

        # Each row increments the num count
        cnt_num += 1

        # PARSE ROW: Start from index 1 (skip the first ID column)
        for idx, cell in enumerate(row[1:], start=1):
            alpha_char = alphabet[cnt_alpha - 1]  # Get the alphabetical character (A, B, C...)
            if idx == 1 and row[1] is not None:  # lemma (first column after ID)
                entry[f"{headers[idx]}_{alpha_char}"] = cell
            else:
                # Remaining cells get lemma and row count
                entry[f"{headers[idx]}_{alpha_char}_{cnt_num}"] = cell

            highest[alpha_char] = max(highest.get(alpha_char, 0), cnt_num)

    # Append the last entry, if any
    if cur_id is not None:
        entry["ID"] = cur_id
        flat_data.append(entry)

    return flat_data, highest

def write_data(data, headers, highest, file_path):
    """
    Write the flattened data to a CSV file.

    :param data: Flattened data entries.
    :param headers: Original headers.
    :param highest: Highest counts for each alphabetical character.
    :param file_path: Path to the output CSV file.
    """
    new_headers = [headers[0]]  # Start with the ID header

    # Create new headers based on highest counts, sorted alphabetically
    for alpha in sorted(highest.keys()):
        new_headers.append(f"{headers[1]}_{alpha}")  # entry_A
        for i in range(1, highest[alpha] + 1):  # 1..X (inclusive)
            for header in headers[2:]:  # SKIP: 1:ID, 2:Entry
                new_headers.append(f"{header}_{alpha}_{i}")  # morpheme_A_1

    # Write the data to the CSV file
    with open(file_path, mode='w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=new_headers)
        writer.writeheader()  # Write the header row

        for entry in data:
            # Prepare entry with all keys initialized to an empty string if not found
            prepared_entry = {key: entry.get(key, "") for key in new_headers}
            writer.writerow(prepared_entry)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Flatten AWS Excel entries and export as CSV.")
    parser.add_argument("-i", "--input", type=str, help="Input file: ASW Excel file", required=False, default="data/AWS.xlsx")
    parser.add_argument("-o", "--output", type=str, help="Output file name", required=False, default="aws_converted.csv")

    args = parser.parse_args()

    # Load, process, and write data
    lines, headers = load_excel(args.input)
    entries, highest = flatten_data(lines, headers)
    write_data(data=entries, headers=headers, highest=highest, file_path=args.output)

    print(highest)

    print(f'"{os.path.basename(__file__)}" finished executing.')